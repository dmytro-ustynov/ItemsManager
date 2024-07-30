import os
from time import time
from decouple import config
import openpyxl
from docx.shared import Mm
from docxtpl import DocxTemplate, InlineImage
from server.app.items.item import FieldNames


class FileExtension:
    xls = "xlsx"
    csv = "csv"
    json = "json"


class FileManager:
    XLS_CONTENT_TYPES = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel')

    @staticmethod
    def create_file(items, extension):
        if extension == FileExtension.xls:
            return FileManager.create_xls(items)
        if extension == FileExtension.csv:
            return FileManager.create_csv(items)
        if extension == FileExtension.json:
            return FileManager.create_json(items)

    @staticmethod
    def create_xls(items: list[dict]):
        # we want name to be the first column
        try:
            field_names = [FieldNames.name]
            for item in items:
                keys = item.keys()
                for field in keys:
                    if field not in field_names:
                        field_names.append(field)
            wb = openpyxl.Workbook()
            sheet = wb.active
            for num, header_name in enumerate(field_names):
                c3 = sheet.cell(row=1, column=num + 1)
                c3.value = header_name
            for item_row_number, item in enumerate(items):
                for column_number, field in enumerate(field_names):
                    value = item.get(field)
                    if value:
                        cell = sheet.cell(row=2 + item_row_number, column=1 + column_number)
                        cell.value = value
            # filename = f"{str(time()).replace('.', '_')}.xls"
            path = FileManager.generate_filepath('xls')
            wb.save(path)
            return True, path
        except Exception as e:
            print(str(e))
            return False, None

    @staticmethod
    def generate_filepath(ext='xls'):
        filename = f"{str(time()).replace('.', '_')}.{ext}"
        path = os.path.join(config("FILE_FOLDER", os.path.join(os.getcwd(), '../data')), filename)
        return path

    @staticmethod
    def create_csv(items):
        False, None

    @staticmethod
    def create_json(items):
        return False, None

    @staticmethod
    def _get_xls_header(wb: openpyxl.Workbook):
        try:
            sheet = wb.get_sheet_by_name(wb.sheetnames[0])
            value = sheet.cell(1, 1).value
            i = 1
            titles = []
            while value:
                value = sheet.cell(1, i).value
                if value:
                    titles.append(value)
                    i += 1
                else:
                    break
            return titles
        except Exception as e:
            print(str(e))
            return []

    @staticmethod
    def validate_input_xls(xls_file=None, path=None):
        """

        :param xls_file:
        :param path:
        :return: (result: boolean, titles:list, rows_count: int)
        """
        wb = None
        try:
            if path is not None and os.path.isfile(path):
                wb = openpyxl.open(path)
            if isinstance(xls_file, openpyxl.Workbook):
                wb = xls_file
            if wb is None:
                raise Exception
        except Exception as e:
            return False, [], 0
        titles = FileManager._get_xls_header(wb)
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        rows = 2
        value = sheet.cell(rows, 1).value
        while value:
            value = sheet.cell(rows, 1).value
            if value:
                rows += 1
            else:
                break
        if titles:
            return True, titles, rows - 2
        else:
            return False, [], 0

    @staticmethod
    def get_data_from_xls(path):
        data = []
        wb = openpyxl.open(path)
        titles = FileManager._get_xls_header(wb)
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])
        i, j = 2, 1
        value = sheet.cell(i, j).value
        while value:
            item = {}
            for j in range(1, len(titles)):
                value = sheet.cell(i, j).value
                if value and titles[j - 1] != FieldNames.ID:
                    item[titles[j - 1]] = value
            data.append(item)
            i += 1
            j = 1
            value = sheet.cell(i, j).value
            if not value:
                break
        return data

    @staticmethod
    def clear_data_folder():
        folder = os.path.join(config("FILE_FOLDER", os.path.join(os.getcwd(), '../data')))
        file_names = list(os.walk(folder))[0][2]
        count = 0
        try:
            for f in file_names:
                if 'xls' in f.split('.')[-1] or 'docx' in f.split('.')[-1]:
                    abs_path = os.path.join(folder, f)
                    if time() - os.path.getmtime(abs_path) > 24 * 3600:
                        os.remove(abs_path)
                        count += 1
            return True, count
        except Exception as e:
            return False, str(e)

    @staticmethod
    def count_data_files():
        folder = os.path.join(config("FILE_FOLDER", os.path.join(os.getcwd(), '../data')))
        file_names = list(os.walk(folder))[0][2]
        count = 0
        try:
            for f in file_names:
                if 'xls' in f.split('.')[-1] or 'docx' in f.split('.')[-1] :
                    count += 1
            return True, count
        except Exception as e:
            return False, str(e)

    @staticmethod
    def generate_tag_file(context):
        template = 'tag_file_tpl.docx'
        filepath = FileManager.generate_filepath('docx')
        doc = DocxTemplate(template)
        img = context.get('qr')
        qr_image = InlineImage(doc, image_descriptor=img, width=Mm(60), height=Mm(60))
        context['qr'] = qr_image
        # context['year'] = context.get('year')
        doc.render(context)
        doc.save(filepath)
        return filepath
